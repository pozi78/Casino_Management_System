import openpyxl
from openpyxl.styles import PatternFill

def get_rgb(fill):
    if fill.start_color.type == 'rgb':
        return fill.start_color.rgb
    if fill.start_color.theme is not None:
        return f"Theme:{fill.start_color.theme}"
    return "None"

wb = openpyxl.load_workbook('Plantilla02.xlsx')
ws = wb.active

print("--- Plantilla02.xlsx Analysis ---")
print(f"Title: {ws.title}")

# Analyze first 20 rows and 10 cols
for row in range(1, 21):
    row_data = []
    for col in range(1, 11): # A-J
        c = ws.cell(row=row, column=col)
        val = str(c.value).strip() if c.value is not None else ""
        fill_rgb = get_rgb(c.fill)
        
        # Identify simplified color
        color_name = "Unknown"
        if fill_rgb == "00000000" or fill_rgb == "None": color_name = "White/None" # Default
        elif fill_rgb == "FFFFFFFF": color_name = "White"
        elif "E" in fill_rgb or "D" in fill_rgb or "C" in fill_rgb: color_name = f"Gray({fill_rgb})" # Roughly gray
        
        if val or color_name != "White/None":
            row_data.append(f"{c.coordinate}:['{val}', {color_name}]")
            
    if row_data:
        print(f"Row {row}: " + ", ".join(row_data))
