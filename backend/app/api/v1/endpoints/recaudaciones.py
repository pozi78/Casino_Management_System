from typing import Any, List, Optional
from datetime import datetime
import shutil
import os
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Query, status, Form, Response
from fastapi.responses import FileResponse
from jose import jwt, JWTError
from pydantic import ValidationError
from app.core.config import settings
from app.schemas.token import TokenPayload
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import Session, joinedload, selectinload
from sqlalchemy.future import select
from app.db.session import get_db
from app.crud.crud_recaudacion import recaudacion, recaudacion_maquina
from app.schemas.recaudacion import (
    Recaudacion as RecaudacionSchema, RecaudacionSummary, RecaudacionCreate, RecaudacionUpdate,
    RecaudacionMaquina as RecaudacionMaquinaSchema, RecaudacionMaquinaUpdate,
    RecaudacionFichero as RecaudacionFicheroSchema
)
from app.models.recaudacion import RecaudacionFichero, Recaudacion, RecaudacionMaquina
from app.api import deps
from app.models.user import Usuario
from app.models.user import Usuario
from app.models.machine import MaquinaExcelMap, Puesto, Maquina
from app.models.salon import Salon
import pandas as pd
from io import BytesIO
import json
from sqlalchemy.dialects.postgresql import insert

router = APIRouter()

@router.get("/last", response_model=Optional[datetime])
async def get_last_recaudacion_date(
    salon_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: Usuario = Depends(deps.get_current_active_user),
) -> Any:
    """
    Get the end date of the last Recaudacion for a Salon.
    Used for auto-filling the start date of the next one.
    """
    # Order by start date descending, get most recent
    recaudaciones = await recaudacion.get_multi(db, skip=0, limit=1, salon_id=salon_id)
    if recaudaciones:
        return recaudaciones[0].fecha_fin
    return None

@router.get("/", response_model=List[RecaudacionSummary])
async def read_recaudaciones(
    skip: int = 0,
    limit: int = 100,
    salon_id: Optional[int] = None,
    db: AsyncSession = Depends(get_db),
    current_user: Usuario = Depends(deps.get_current_active_user),
) -> Any:
    """
    Retrieve recaudaciones.
    """
    recaudaciones = await recaudacion.get_multi(db, skip=skip, limit=limit, salon_id=salon_id)
    recaudaciones = await recaudacion.get_multi(db, skip=skip, limit=limit, salon_id=salon_id)
    return recaudaciones

from pydantic import BaseModel
class RecaudacionMetadata(BaseModel):
    is_normalized: bool
    salon_id: Optional[int] = None
    fecha_inicio: Optional[datetime] = None
    fecha_fin: Optional[datetime] = None
    error: Optional[str] = None

@router.post("/parse-metadata", response_model=RecaudacionMetadata)
async def parse_recaudacion_metadata(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    current_user: Usuario = Depends(deps.get_current_active_user),
) -> Any:
    """
    Parse Excel file to extract metadata for form pre-filling.
    """
    contents = await file.read()
    try:
        df = pd.read_excel(BytesIO(contents), header=None)
    except Exception as e:
         return {"is_normalized": False, "error": f"Error parsing excel: {str(e)}"}

    # Check for Normalization (Version 1.0 in D3)
    is_normalized = False
    try:
        if df.shape[0] > 3 and df.shape[1] > 4:
            val = df.iloc[2, 3] # D3
            if pd.notna(val) and isinstance(val, str) and "VERSION" in val.upper():
                is_normalized = True
    except:
        pass
    
    if not is_normalized:
        return {"is_normalized": False}

    # Extract Data if Normalized
    # A1: "SALON: {Name}" (Row 0, Col 0)
    # E1: End Date (Row 0, Col 4) -> Current Recaudacion Date -> Form End Date
    # E2: Start Date (Row 1, Col 4) -> Previous Recaudacion Date -> Form Start Date
    
    res = {"is_normalized": True}
    
    # Salon
    try:
        # Export format might be "SALON: Nombre" in A1 OR "SALON" in A1 and Name in B1
        cell_val = df.iloc[0, 0] # A1
        print(f"DEBUG: Parsing metadata. A1 value: '{cell_val}'")
        
        salon_name = None
        
        if pd.notna(cell_val) and isinstance(cell_val, str):
            clean_a1 = cell_val.strip().upper()
            if clean_a1 == "SALON" or clean_a1 == "SALON:":
                # Name is in B1
                val_b1 = df.iloc[0, 1]
                print(f"DEBUG: A1 is label. Checking B1: '{val_b1}'")
                if pd.notna(val_b1) and isinstance(val_b1, str):
                    salon_name = val_b1.strip()
            elif "SALON:" in clean_a1:
                 # Name is in A1
                 salon_name = cell_val.replace("SALON:", "").replace("Salon:", "").strip()
            else:
                 # Assuming A1 is the name directly? Or B1?
                 # Let's try B1 just in case
                 pass
        
        if not salon_name:
             # Fallback: Try B1 directly if A1 failed to provide a name
             val_b1 = df.iloc[0, 1]
             if pd.notna(val_b1) and isinstance(val_b1, str):
                 salon_name = val_b1.strip()

        if salon_name:
            print(f"DEBUG: Final extracted salon name: '{salon_name}'")
            
            # Find Salon by Name (Try exact first, then contained?)
            stmt = select(Salon).where(Salon.nombre.ilike(salon_name))
            salon = (await db.execute(stmt)).scalars().first()
            
            if not salon:
                print(f"DEBUG: Exact match failed for '{salon_name}'. Trying containment.")
                stmt = select(Salon).where(Salon.activo == True)
                all_salons = (await db.execute(stmt)).scalars().all()
                for s in all_salons:
                    # Robust fuzzy match
                    # Check if DB name is in CSV name OR CSV name is in DB Name
                    if s.nombre.lower() in salon_name.lower() or salon_name.lower() in s.nombre.lower():
                        salon = s
                        print(f"DEBUG: Fuzzy match found: '{s.nombre}'")
                        break
            
            if salon:
                res["salon_id"] = salon.id
                print(f"DEBUG: Matched Salon ID: {salon.id}")
            else:
                 print(f"DEBUG: No salon matched for '{salon_name}'")
        else:
            print("DEBUG: Could not extract any salon name string.")
    except Exception as e:
        print(f"Error parsing salon: {e}")

    # Dates
    def parse_date(val):
        if isinstance(val, datetime):
            return val
        if isinstance(val, str):
            try:
                return datetime.fromisoformat(val)
            except:
                pass
        return None

    try:
        # E1 (Row 0, Col 4) is DATE OF THIS RECAUDACION (End Date)
        # E2 (Row 1, Col 4) is DATE OF PREVIOUS (Start Date)
        
        val_e1 = df.iloc[0, 4] # E1
        val_e2 = df.iloc[1, 4] # E2
        
        # Current Form Logic:
        # Fecha Inicio = Start of period (E2)
        # Fecha Fin = End of period (E1)
        
        if pd.notna(val_e2):
             res["fecha_inicio"] = val_e2
             
        if pd.notna(val_e1):
             res["fecha_fin"] = val_e1
            
    except Exception as e:
        print(f"Error parsing dates: {e}")

    return res

@router.post("/", response_model=RecaudacionSchema)
async def create_recaudacion(
    recaudacion_in: RecaudacionCreate,
    db: AsyncSession = Depends(get_db),
    current_user: Usuario = Depends(deps.get_current_active_user),
) -> Any:
    """
    Create new recaudacion and auto-generate machine details.
    """
    recaudacion_obj = await recaudacion.create_with_initial_details(db, obj_in=recaudacion_in)
    return recaudacion_obj

@router.get("/{id}", response_model=RecaudacionSchema)
async def read_recaudacion(
    id: int,
    db: AsyncSession = Depends(get_db),
    current_user: Usuario = Depends(deps.get_current_active_user),
) -> Any:
    """
    Get recaudacion by ID.
    """
    recaudacion_obj = await recaudacion.get(db, id=id)
    if not recaudacion_obj:
        raise HTTPException(status_code=404, detail="Recaudacion not found")
    return recaudacion_obj

@router.put("/{id}", response_model=RecaudacionSchema)
async def update_recaudacion(
    id: int,
    recaudacion_in: RecaudacionUpdate,
    db: AsyncSession = Depends(get_db),
    current_user: Usuario = Depends(deps.get_current_active_user),
) -> Any:
    """
    Update recaudacion header.
    """
    recaudacion_obj = await recaudacion.get(db, id=id)
    if not recaudacion_obj:
        raise HTTPException(status_code=404, detail="Recaudacion not found")
    recaudacion_updated = await recaudacion.update(db, db_obj=recaudacion_obj, obj_in=recaudacion_in)
    
    # Recalculate Tasa Diferencia if total_tasas changed (or always to be safe)
    # Check if Dates Changed -> Recalculate Estimated Taxes
    dates_changed = False
    in_dump = recaudacion_in.model_dump(exclude_unset=True)
    if 'fecha_inicio' in in_dump or 'fecha_fin' in in_dump:
        await recalculate_estimated_taxes(db, id)
        await db.refresh(recaudacion_obj) # Refresh to get new details if needed? Logic updates details directly.
        dates_changed = True

    if 'total_tasas' in in_dump or dates_changed:
         await recalculate_tasa_diferencia(db, id)
         
    return recaudacion_updated

async def recalculate_estimated_taxes(db: AsyncSession, recaudacion_id: int):
    # Fetch Recaudacion with details and machine info
    stmt = select(Recaudacion).options(
        selectinload(Recaudacion.detalles).options(
            selectinload(RecaudacionMaquina.maquina).selectinload(Maquina.tipo_maquina),
            selectinload(RecaudacionMaquina.puesto)
        )
    ).where(Recaudacion.id == recaudacion_id)
    rec = (await db.execute(stmt)).scalars().first()
    
    if not rec or not rec.detalles:
        return

    # Calculate Days
    if not rec.fecha_inicio or not rec.fecha_fin:
        return
        
    days_diff = (rec.fecha_fin - rec.fecha_inicio).days
    if days_diff < 0: days_diff = 0
    
    for d in rec.detalles:
        # Re-evaluate weekly rate logic (Duplicate from CRUD, ideally refactor to helper)
        tasa_base = 0
        
        # Determine Weekly Rate
        weekly_rate = 0
        if d.puesto and d.puesto.tasa_semanal:
            weekly_rate = d.puesto.tasa_semanal
        elif d.maquina:
             if d.maquina.tasa_semanal_override:
                 weekly_rate = d.maquina.tasa_semanal_override
             elif d.maquina.tipo_maquina and d.maquina.tipo_maquina.tasa_semanal_orientativa:
                 weekly_rate = d.maquina.tipo_maquina.tasa_semanal_orientativa
        
        # Calculate
        if float(weekly_rate) > 0 and days_diff > 0:
             daily_rate = float(weekly_rate) / 7.0
             tasa_base = daily_rate * days_diff
        
        d.tasa_estimada = tasa_base
        # tasa_final will be updated by recalculate_tasa_diferencia next
        db.add(d)
        
    await db.commit()

async def recalculate_tasa_diferencia(db: AsyncSession, recaudacion_id: int):
    # 1. Get Recaudacion + Details
    # Need to load details to iterate
    stmt = select(Recaudacion).options(selectinload(Recaudacion.detalles)).where(Recaudacion.id == recaudacion_id)
    rec = (await db.execute(stmt)).scalars().first()
    if not rec or not rec.detalles:
        return

    # 2. Totals
    total_real = float(rec.total_tasas or 0)
    total_calculada = sum(float(d.tasa_estimada or 0) for d in rec.detalles)
    
    diff = total_real - total_calculada
    
    # 3. Distribute
    # If total_calculada is 0, we can't distribute roughly. Avoid division by zero.
    # If 0, maybe assign 0 to difference?
    
    for d in rec.detalles:
        t_calc = float(d.tasa_estimada or 0)
        
        if total_calculada != 0:
            weight = t_calc / total_calculada
            d.tasa_diferencia = round(diff * weight, 4)
        else:
            d.tasa_diferencia = 0
            
        # Update Final
        # Tasa Final = Calculada + Diferencia + Ajuste
        d.tasa_final = float(d.tasa_estimada or 0) + float(d.tasa_diferencia or 0) + float(d.ajuste or 0)
        
        db.add(d)
        
    await db.commit()

# --- Detail Endpoints ---

@router.put("/details/{detail_id}", response_model=RecaudacionMaquinaSchema)
async def update_recaudacion_detail(
    detail_id: int,
    detail_in: RecaudacionMaquinaUpdate,
    db: AsyncSession = Depends(get_db),
    current_user: Usuario = Depends(deps.get_current_active_user),
) -> Any:
    """
    Update a specific machine line (retirada, cajon, etc).
    """
    detail_obj = await recaudacion_maquina.get(db, id=detail_id)
    if not detail_obj:
        raise HTTPException(status_code=404, detail="Recaudacion Detail not found")
    
    # We could check here if Recaudacion is closed/locked if we had that logic
    
    updated_detail = await recaudacion_maquina.update(db, db_obj=detail_obj, obj_in=detail_in)
    return updated_detail

@router.delete("/details/{detail_id}", response_model=RecaudacionMaquinaSchema)
async def delete_recaudacion_detail(
    detail_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: Usuario = Depends(deps.get_current_active_user),
) -> Any:
    """
    Delete a specific machine line.
    """
    detail_obj = await recaudacion_maquina.get(db, id=detail_id)
    if not detail_obj:
        raise HTTPException(status_code=404, detail="Recaudacion Detail not found")
    
    # We could check here if Recaudacion is closed/locked
    recaudacion_obj = await recaudacion.get(db, id=detail_obj.recaudacion_id)
    if recaudacion_obj and recaudacion_obj.bloqueada:
         raise HTTPException(status_code=400, detail="Cannot delete details from a locked recaudacion")

    detail_deleted = await recaudacion_maquina.remove(db, id=detail_id)
    
    # Expunge parent to force reload of details from scratch
    db.expunge(recaudacion_obj)
    
    # Recalculate differences for remaining items
    await recalculate_tasa_diferencia(db, recaudacion_obj.id)
    
    return detail_deleted

@router.delete("/{id}", response_model=RecaudacionSchema)
async def delete_recaudacion(
    id: int,
    db: AsyncSession = Depends(get_db),
    current_user: Usuario = Depends(deps.get_current_active_user),
) -> Any:
    """
    Delete recaudacion.
    """
    recaudacion_obj = await recaudacion.get(db, id=id)
    if not recaudacion_obj:
        raise HTTPException(status_code=404, detail="Recaudacion not found")
    recaudacion_deleted = await recaudacion.remove(db, id=id)
    return recaudacion_deleted

# --- File Handling ---

UPLOAD_DIR = "/opt/CasinosSM/backend/uploads/recaudaciones"

import random
import string
def generate_unique_filename(filename: str) -> str:
    code = ''.join(random.choices(string.ascii_uppercase + string.digits, k=10))
    if "." in filename:
        parts = filename.rsplit(".", 1)
        return f"{parts[0]}.{code}.{parts[1]}"
    return f"{filename}.{code}"

@router.post("/{id}/files", response_model=RecaudacionFicheroSchema)
async def upload_recaudacion_file(
    id: int,
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    current_user: Usuario = Depends(deps.get_current_active_user),
) -> Any:
    # Check if recaudacion exists
    rec = await recaudacion.get(db, id=id)
    if not rec:
        raise HTTPException(status_code=404, detail="Recaudacion not found")
        
    # Ensure directory exists
    rec_dir = os.path.join(settings.UPLOAD_DIR, "recaudaciones", str(id))
    os.makedirs(rec_dir, exist_ok=True)
    
    # Unique Filename
    unique_name = generate_unique_filename(file.filename)
    file_path = os.path.join(rec_dir, unique_name)
    
    with open(file_path, "wb") as buffer:
        shutil.copyfileobj(file.file, buffer)
        
    # Create DB Entry
    db_file = RecaudacionFichero(
        recaudacion_id=id,
        file_path=file_path,
        filename=file.filename, # Keep Original Name
        content_type=file.content_type,
        created_at=datetime.now()
    )
    db.add(db_file)
    await db.commit()
    await db.refresh(db_file)
    return db_file
    
# ... And update import_recaudacion_excel ...

@router.delete("/{id}/files/{file_id}")
async def delete_recaudacion_file(
    id: int,
    file_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: Usuario = Depends(deps.get_current_active_user),
) -> Any:
    stmt = select(RecaudacionFichero).where(RecaudacionFichero.id == file_id, RecaudacionFichero.recaudacion_id == id)
    result = await db.execute(stmt)
    db_file = result.scalars().first()
    
    if not db_file:
         raise HTTPException(status_code=404, detail="File not found")
         
    # Delete from Disk
    if os.path.exists(db_file.file_path):
        os.remove(db_file.file_path)
        
    await db.delete(db_file)
    await db.commit()
    return {"status": "success"}

@router.get("/files/{file_id}/content")
async def get_file_content(
    file_id: int,
    db: AsyncSession = Depends(get_db),
    token: Optional[str] = Query(None),
) -> Any:
    # Validate Token from Query Param
    if not token:
        raise HTTPException(status_code=401, detail="Not authenticated")
    try:
        payload = jwt.decode(token, settings.SECRET_KEY, algorithms=[settings.ALGORITHM])
        TokenPayload(**payload)
    except (JWTError, ValidationError):
        raise HTTPException(status_code=403, detail="Could not validate credentials")

    stmt = select(RecaudacionFichero).where(RecaudacionFichero.id == file_id)
    result = await db.execute(stmt)
    db_file = result.scalars().first()
    
    if not db_file or not os.path.exists(db_file.file_path):
        raise HTTPException(status_code=404, detail="File not found")
        
    return FileResponse(
        db_file.file_path, 
        media_type=db_file.content_type or "application/octet-stream",
        filename=db_file.filename,
        content_disposition_type="inline"
    )

@router.post("/{id}/analyze-excel")
async def analyze_recaudacion_excel(
    id: int,
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    current_user: Usuario = Depends(deps.get_current_active_user),
) -> Any:
    rec = await recaudacion.get(db=db, id=id)
    if not rec: raise HTTPException(404, "Recaudacion not found")
    
    # 1. Parse Excel Names
    contents = await file.read()
    try:
        df = pd.read_excel(BytesIO(contents), header=None)
    except Exception as e:
         raise HTTPException(400, f"Error parsing excel: {e}")

    excel_names = set()
    # Basic heuristic: Check first 20 columns, first 50 rows
    for col in range(min(20, df.shape[1])):
        for row in range(min(100, df.shape[0])):
            val = df.iloc[row, col]
            if pd.notna(val) and isinstance(val, str) and len(val) > 2:
                 # Filter out short strings or likely non-names
                 excel_names.add(val.strip().upper())

    # Check for Normalization (Version 1.0)
    is_normalized = False
    try:
        # Check D3 (Row 2, Col 3) for "VERSION"
        if df.shape[0] > 3 and df.shape[1] > 4:
            val = df.iloc[2, 3]
            if pd.notna(val) and isinstance(val, str) and "VERSION" in val.upper():
                is_normalized = True
    except:
        pass
                 
    return await _process_analysis_result(excel_names, rec.salon_id, db, is_normalized)

async def _process_analysis_result(excel_names: set, salon_id: int, db: AsyncSession, is_normalized: bool = False):
    # 2. Get Existing Mappings
    stmt = select(MaquinaExcelMap).where(MaquinaExcelMap.salon_id == salon_id)
    existing_maps = {m.excel_nombre.upper(): m for m in (await db.execute(stmt)).scalars().all()}
    
    # 3. Get All Puestos in Salon
    stmt_p = select(Puesto, Maquina).join(Maquina).where(Maquina.salon_id == salon_id, Puesto.activo == True).order_by(Maquina.nombre, Puesto.numero_puesto)
    puestos_res = (await db.execute(stmt_p)).all()
    
    puestos_list = []
    for p, m in puestos_res:
         label = f"{m.nombre} - Puesto {p.numero_puesto}"
         if p.descripcion: label += f" ({p.descripcion})"
         puestos_list.append({"id": p.id, "name": label})
         
    # Build Result
    mappings_result = []
    excluded_keywords = ["MAQUINA", "TOTAL", "PAGO MANUAL", "RETIRADA", "CAJON", "IMPUESTOS", "DPS"]
    for name in sorted(list(excel_names)):
        # Skip likely headers
        if any(k in name for k in excluded_keywords): continue
        
        existing = existing_maps.get(name)
        mappings_result.append({
            "excel_name": name,
            "mapped_puesto_id": existing.puesto_id if existing else None,
            "is_ignored": existing.is_ignored if existing else False
        })
        
    return {"mappings": mappings_result, "puestos": puestos_list, "is_normalized": is_normalized}


@router.post("/{id}/files/{file_id}/analyze")
async def analyze_recaudacion_file(
    id: int,
    file_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: Usuario = Depends(deps.get_current_active_user),
) -> Any:
    # 1. Get Recaudacion & File
    rec = await recaudacion.get(db=db, id=id)
    if not rec: raise HTTPException(404, "Recaudacion not found")
    
    stmt = select(RecaudacionFichero).where(RecaudacionFichero.id == file_id, RecaudacionFichero.recaudacion_id == id)
    db_file = (await db.execute(stmt)).scalars().first()
    
    if not db_file or not os.path.exists(db_file.file_path):
        raise HTTPException(404, "File not found")
        
    # 2. Read & Parse
    try:
        with open(db_file.file_path, "rb") as f:
            contents = f.read()
        df = pd.read_excel(BytesIO(contents), header=None)
    except Exception as e:
         raise HTTPException(400, f"Error parsing excel: {e}")

    excel_names = set()
    for col in range(min(20, df.shape[1])):
        for row in range(min(100, df.shape[0])):
            val = df.iloc[row, col]
            if pd.notna(val) and isinstance(val, str) and len(val) > 2:
                 excel_names.add(val.strip().upper())

                 excel_names.add(val.strip().upper())

    # Check for Normalization (Version 1.0)
    is_normalized = False
    try:
        if df.shape[0] > 3 and df.shape[1] > 4:
            val = df.iloc[2, 3]
            if pd.notna(val) and isinstance(val, str) and "VERSION" in val.upper():
                is_normalized = True
    except:
        pass

    return await _process_analysis_result(excel_names, rec.salon_id, db, is_normalized)


@router.post("/{id}/import-excel")
async def import_recaudacion_excel(
    id: int,
    mappings_str: Optional[str] = Form(None),
    file: Optional[UploadFile] = File(None),
    file_id: Optional[int] = Form(None),
    db: AsyncSession = Depends(get_db),
    current_user: Usuario = Depends(deps.get_current_active_user),
) -> Any:
    import sys
    print(f"DEBUG: import_recaudacion_excel called. id={id}, file_id={file_id}, file={file}, mappings_str={mappings_str}")
    sys.stdout.flush()
    
    # 1. Verify Recaudacion
    rec = await recaudacion.get(db=db, id=id)
    if not rec:
        raise HTTPException(status_code=404, detail="Recaudacion not found")
        
    # Check Inputs
    if not file and not file_id:
         print(f"DEBUG: Missing inputs. file={file}, file_id={file_id}")
         sys.stdout.flush()
         raise HTTPException(status_code=400, detail=f"Must provide either file or file_id. Received file={file}, file_id={file_id}")

    # 1.5 Update Mappings if provided
    if mappings_str:
        try:
            new_mappings = json.loads(mappings_str)
            for name, pid in new_mappings.items():
                name = name.strip().upper()
                # pid can be None, Int, or -1 (Ignore)
                
                is_ignored = False
                puesto_id = None
                
                if pid == -1:
                    is_ignored = True
                elif pid:
                    puesto_id = int(pid)
                
                # Check exist
                stmt = select(MaquinaExcelMap).where(
                    MaquinaExcelMap.salon_id == rec.salon_id, 
                    MaquinaExcelMap.excel_nombre == name
                )
                existing = (await db.execute(stmt)).scalars().first()
                if existing:
                    existing.puesto_id = puesto_id
                    existing.maquina_id = None
                    existing.is_ignored = is_ignored
                    db.add(existing)
                else:
                    new_map = MaquinaExcelMap(
                        salon_id=rec.salon_id,
                        excel_nombre=name,
                        puesto_id=puesto_id,
                        maquina_id=None,
                        is_ignored=is_ignored
                    )
                    db.add(new_map)
            await db.commit()
        except Exception as e:
            print(f"Error updating mappings: {e}")
            
    # 2. Get File Content & Save if new
    contents = b""
    filename = ""
    
    if file_id:
        stmt = select(RecaudacionFichero).where(RecaudacionFichero.id == file_id, RecaudacionFichero.recaudacion_id == id)
        db_file = (await db.execute(stmt)).scalars().first()
        if not db_file or not os.path.exists(db_file.file_path):
             raise HTTPException(404, "File not found")
        with open(db_file.file_path, "rb") as f:
            contents = f.read()
        filename = db_file.filename
    else:
        # New Upload
        contents = await file.read()
        filename = file.filename
        
        # Save File
        rec_dir = os.path.join(settings.UPLOAD_DIR, "recaudaciones", str(id))
        os.makedirs(rec_dir, exist_ok=True)
        unique_name = generate_unique_filename(filename)
        file_path = os.path.join(rec_dir, unique_name)
        
        with open(file_path, "wb") as f:
            f.write(contents)
            
        new_file = RecaudacionFichero(
            recaudacion_id=id,
            file_path=file_path,
            filename=filename,
            content_type=file.content_type,
            created_at=datetime.now()
        )
        db.add(new_file)
        await db.commit()

    # 3. Process Excel
    try:
        df = pd.read_excel(BytesIO(contents), header=None)
    except Exception as e:
         raise HTTPException(400, f"Error parsing excel: {e}")

    # Detect Normalization
    is_normalized = False
    try:
        if df.shape[0] > 3 and df.shape[1] > 4:
            val = df.iloc[2, 3]
            if pd.notna(val) and isinstance(val, str) and "VERSION" in val.upper():
                is_normalized = True
    except:
        pass

    updated_count = 0
    
    # 4. Prepare Maps for Import
    # Fetch Mappings
    stmt_map = select(MaquinaExcelMap).where(MaquinaExcelMap.salon_id == rec.salon_id)
    name_map = {m.excel_nombre.upper(): m for m in (await db.execute(stmt_map)).scalars().all()}
    
    # Fetch Details for this Recaudacion to update them
    # We assume details already satisfy uniqueness by puesto/maquina for this recaudacion
    stmt_details = select(RecaudacionMaquina).options(
        joinedload(RecaudacionMaquina.maquina),
        joinedload(RecaudacionMaquina.puesto)
    ).where(RecaudacionMaquina.recaudacion_id == id)
    current_details = (await db.execute(stmt_details)).scalars().all()
    
    details_by_puesto = {d.puesto_id: d for d in current_details if d.puesto_id}
    details_by_maquina = {d.maquina_id: d for d in current_details if d.maquina_id}
    # Fallback map for normalized imports (Direct Name Match)
    details_by_name = {}
    for d in current_details:
        if d.maquina and d.maquina.nombre:
            key_name = d.maquina.nombre.strip()
            # If multi-post (has Puesto), append suffixes compatible with Export Logic
            if d.puesto:
                p_desc = d.puesto.descripcion
                p_numero = d.puesto.numero_puesto
                puesto_str = f" - {p_desc}" if p_desc else (f" - PUESTO {p_numero}" if p_numero else "")
                key_name = f"{key_name}{puesto_str}"
            
            details_by_name[key_name.upper()] = d

    # 5. Iterate Rows
    for r_idx, row in df.iterrows():
        
        detail = None
        
        if is_normalized:
            # Normalized Format:
            # Start at Row 12 (Index 12)
            if r_idx < 12: continue
            
            # Col 0: Name, Col 1: Retirada, Col 2: Cajon, Col 3: Manual, Col 4: Ajuste
            if len(row) < 5: continue
            
            raw_name = row[0]
            if pd.isna(raw_name) or not isinstance(raw_name, str): continue
            
            # For normalized, we expect exact name matches usually, OR we use the map if they were mapped?
            # The prompt says: "use the same mapping system ... if normalized make automatic import taking into account version".
            # Automatic import implies we trust the names or rely on the mapping system?
            # If it's normalized (exported by us), the names should match exactly with `d["MAQUINA"]` which is `f"{m_nombre}{puesto_str}"`.
            # HOWEVER, `name_map` is keyed by `excel_nombre`. If we use the map, we are safe.
            # analyze_excel adds names to the map.
            
            clean_name = raw_name.strip().upper()
            
            if clean_name in name_map:
                print(f"DEBUG: Found in name_map: '{clean_name}'")
                mapping = name_map[clean_name]
                if mapping.puesto_id:
                     detail = details_by_puesto.get(mapping.puesto_id)
                elif mapping.maquina_id:
                     detail = details_by_maquina.get(mapping.maquina_id)
            else:
                 # Helper: Try direct match with machine name
                 # This covers the case where the normalized file uses valid system names but they haven't been "Mapped" in MaquinaExcelMap yet.
                 print(f"DEBUG: Not in map. Checking details_by_name for: '{clean_name}'")
                 if clean_name in details_by_name:
                     print(f"DEBUG: Found exact match in details_by_name for '{clean_name}'")
                     detail = details_by_name[clean_name]
                 else:
                     # Fallback 2: Check containment?
                     # Normalized export might include Puesto number "Name - P1"
                     for k, d in details_by_name.items():
                         if k in clean_name or clean_name in k:
                              print(f"DEBUG: Fuzzy match in details_by_name: '{clean_name}' ~= '{k}'")
                              detail = d
                              break
                     if not detail:
                          print(f"DEBUG: Failed to match '{clean_name}'")

            if detail:
                def get_val(idx):
                    v = row[idx] if idx < len(row) else 0
                    return float(v) if pd.notna(v) and isinstance(v, (int, float)) else 0
                
                detail.retirada_efectivo = get_val(1)
                detail.cajon = get_val(2)
                detail.pago_manual = get_val(3)
                detail.ajuste = get_val(4)
                
                db.add(detail)
                updated_count += 1
                
        else:
            # Legacy Format
            # Check Column 1 (Index 1) for Name
            if len(row) < 2: continue
            
            raw_name = row[1]
            if pd.isna(raw_name) or not isinstance(raw_name, str):
                continue
                
            clean_name = raw_name.strip().upper()
            
            # Check Map
            if clean_name in name_map:
                mapping = name_map[clean_name]
                
                detail = None
                if mapping.puesto_id:
                    detail = details_by_puesto.get(mapping.puesto_id)
                elif mapping.maquina_id:
                    detail = details_by_maquina.get(mapping.maquina_id)
                
                if detail:
                    # Extract Values. Col 5: Retirada, 6: Cajon, 7: Pago Manual
                    def get_val(idx):
                        v = row[idx] if idx < len(row) else 0
                        return float(v) if pd.notna(v) and isinstance(v, (int, float)) else 0
                    
                    detail.retirada_efectivo = get_val(5)
                    detail.cajon = get_val(6)
                    detail.pago_manual = get_val(7)
                    
                    db.add(detail)
                    updated_count += 1

    # 6. Extract Totals (Impuestos, DPS)
    # Based on analysis: IMPUESTOS at Col 3, Value at Col 5.
    found_totals = False
    
    if not is_normalized:
        for r_idx, row in df.iterrows():
            if len(row) < 6: continue
            
            label_col = 3
            val_col = 5
            
            label = row[label_col]
            if pd.isna(label) or not isinstance(label, str):
                continue
                
            label = label.upper()
            
            if "IMPUESTOS" in label:
                val = row[val_col]
                if pd.notna(val) and isinstance(val, (int, float)):
                    rec.total_tasas = float(val)
                    found_totals = True
                    
            if "DPS" in label:
                val = row[val_col]
                if pd.notna(val) and isinstance(val, (int, float)):
                    rec.depositos = float(val)
    else:
        # Normalized files extraction (Rows 3-10 contain summary)
        # Row 6 (Index 5): TOTAL TASAS (Col 1/B)
        # Row 8 (Index 7): DEPOSITOS (Col 1/B)
        # Row 9 (Index 8): OTROS CONCEPTOS (Col 1/B)
        print("DEBUG: Extracting totals for normalized file")
        
        def get_val_at(r, c):
             if r < df.shape[0] and c < df.shape[1]:
                 val = df.iloc[r, c]
                 if pd.notna(val) and isinstance(val, (int, float)):
                     return float(val)
             return 0.0

        # Row indices are 0-based
        rec.total_tasas = get_val_at(5, 1)    # Row 6
        rec.depositos = get_val_at(7, 1)      # Row 8
        rec.otros_conceptos = get_val_at(8, 1) # Row 9
        
        found_totals = True
        print(f"DEBUG: Extracted Totals - Tasas: {rec.total_tasas}, Dep: {rec.depositos}, Otros: {rec.otros_conceptos}")
                
    if found_totals:
        db.add(rec)
        
    await db.commit()
    
    # Recalculate Tasa Diferencia after import
    await recalculate_tasa_diferencia(db, id)
    
    return {"status": "ok", "updated": updated_count, "mappings_updated": True if mappings_str else False}


@router.get("/{id}/export-excel")
async def export_recaudacion_excel(
    id: int,
    db: AsyncSession = Depends(get_db),
    current_user: Usuario = Depends(deps.get_current_active_user),
) -> Any:
    # 1. Fetch Recaudacion
    # Eagerly load 'salon' to avoid MissingGreenlet error during async access
    stmt = select(Recaudacion).options(selectinload(Recaudacion.salon)).where(Recaudacion.id == id)
    rec = (await db.execute(stmt)).scalars().first()
    if not rec:
        raise HTTPException(404, "Recaudacion not found")
        
    # 2. Fetch Previous Recaudacion for Header Info
    # Get the most recent one before this one for the same salon
    stmt_prev = select(Recaudacion).where(
        Recaudacion.salon_id == rec.salon_id,
        Recaudacion.fecha_inicio < rec.fecha_inicio
    ).order_by(Recaudacion.fecha_inicio.desc()).limit(1)
    
    prev_rec = (await db.execute(stmt_prev)).scalars().first()
    
    prev_date_str = prev_rec.fecha_fin.strftime('%d/%m/%Y') if prev_rec and prev_rec.fecha_fin else "N/A"
    days_diff = (rec.fecha_inicio - prev_rec.fecha_fin).days if prev_rec and prev_rec.fecha_fin and rec.fecha_inicio else 0
    
    # 3. Fetch Data
    stmt_data = select(
        RecaudacionMaquina, 
        Maquina.nombre, 
        Maquina.numero_serie,
        Puesto.numero_puesto, 
        Puesto.descripcion, 
        Maquina.grupo_id
    ).outerjoin(Maquina, RecaudacionMaquina.maquina_id == Maquina.id)\
     .outerjoin(Puesto, RecaudacionMaquina.puesto_id == Puesto.id)\
     .where(RecaudacionMaquina.recaudacion_id == id)\
     .order_by(Maquina.nombre, Puesto.numero_puesto)
     
    results = (await db.execute(stmt_data)).all()
    
    # Sort: Multipuesto First, then Name, then Puesto
    from collections import Counter
    # Count occurrences by machine name
    name_counts = Counter(r[1] for r in results) # index 1 is Maquina.nombre

    # Sort key function
    def sort_key(row):
        m_name = row[1]
        p_numero = row[3] # Puesto.numero_puesto
        is_multi = name_counts[m_name] > 1
        # Tuple comparison:
        # 0. not is_multi (False < True, so reverse this logic or use not: multipuesto should come first (True/1))
        # Let's say: 0 if Multi, 1 if Mono. 
        type_rank = 0 if is_multi else 1
        return (type_rank, m_name, p_numero or 0)
        
    results.sort(key=sort_key)
    
    data = []
    for row in results:
        det, m_nombre, m_serie, p_numero, p_desc, g_id = row
        puesto_str = f" - {p_desc}" if p_desc else (f" - PUESTO {p_numero}" if p_numero else "")
        maquina_display = f"{m_nombre}{puesto_str}"
        
        data.append({
            "MAQUINA": maquina_display,
            "RETIRADA": det.retirada_efectivo or 0,
            "CAJON": det.cajon or 0,
            "PAGO MANUAL": det.pago_manual or 0,
            "AJUSTE": det.ajuste or 0,
            "TASA_EST": det.tasa_estimada or 0,
            "raw_name": m_nombre # For grouping
        })
        
    # 4. Excel Generation
    output = BytesIO()
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Protection as CellProtection, Border, Side, Alignment
    from openpyxl.formatting.rule import CellIsRule
    from openpyxl.styles.differential import DifferentialStyle
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Recaudación"
    
    # 4. Excel Generation
    output = BytesIO()
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, Protection as CellProtection
    
    # Sheet Name: DD - MMM - AAAA (Spanish)
    months_es = {
        1: "ene", 2: "feb", 3: "mar", 4: "abr", 5: "may", 6: "jun",
        7: "jul", 8: "ago", 9: "sep", 10: "oct", 11: "nov", 12: "dic"
    }
    day_str = rec.fecha_fin.strftime('%d')
    month_str = months_es[rec.fecha_fin.month]
    year_str = rec.fecha_fin.strftime('%Y')
    ws.title = f"{day_str} - {month_str} - {year_str}"
    
    # Enable Sheet Protection (default is locked for all cells)
    ws.protection.sheet = True
    
    # Styles
    bold_font = Font(bold=True)
    red_static_font = Font(color="FF0000", bold=True)
    # Darker Gray for editable cells (User requested "un poco más oscuro")
    # Previous was EEEEEE. Let's try CCCCCC or DDDDDD.
    gray_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid") # LightGray
    
    # --- Header Info (Rows 1-2) ---
    ws['A1'] = f"SALON: {rec.salon.nombre if rec.salon else ''}"
    ws['A1'].font = bold_font
    
    ws['D1'] = "RECAUDACION"
    ws['D1'].font = bold_font
    
    c = ws['E1']
    c.value = rec.fecha_fin.strftime('%d/%m/%Y')
    c.font = bold_font
    c.protection = CellProtection(locked=False) # Editable
    c.fill = gray_fill

    ws['D2'] = "ULTIMA RECAUDACION"
    ws['D2'].font = bold_font
    
    c = ws['E2']
    c.value = prev_date_str
    c.font = bold_font
    c.protection = CellProtection(locked=False) # Editable
    c.fill = gray_fill
    
    ws['F2'] = "DIAS"
    ws['F2'].font = bold_font
    
    # G2 Formula: Days diff. Try to use formula if possible, or static value but template suggests formula.
    # DATEDIF is tricky in Excel vs OpenPyol depending on locale. Simple subtraction is better: =E1-E2
    ws['G2'] = f"=E1-E2"
    ws['G2'].font = bold_font
    
    # --- Styling & Configuration ---
    # Colors: User wants Inverted relative to Plantilla02 (where Editable was Gray).
    # So: Editable -> White (FFFFFF), Locked -> Gray (D3D3D3).
    gray_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    
    bold_font = Font(bold=True)
    header_font = Font(bold=True, size=12)
    money_fmt = '#,##0.00'
    
    start_row = 13
    end_row = start_row + len(data) - 1 if len(data) > 0 else start_row
    t_row = end_row + 1 # Row for Totals
    
    # 1. Base Fill: Apply Gray to the entire active area (Structure = Gray)
    # Range: A1 to H{t_row}.
    for r in range(1, t_row + 1):
        for c in range(1, 9): # A-H
            ws.cell(row=r, column=c).fill = gray_fill

    # --- Header (Rows 1-2) ---
    # A1: SALON (Label) - Font 16
    ws['A1'] = "SALON"
    ws['A1'].font = Font(bold=True, size=16)
    
    # B1: Salon Name - Font 16
    ws['B1'] = rec.salon.nombre if rec.salon else ""
    ws['B1'].font = Font(bold=True, size=16)
    
    # D1: RECAUDACION - Font 16
    ws['D1'] = "RECAUDACION"
    ws['D1'].font = Font(bold=True, size=16)
    
    # E1: Current Date - Font 16, Editable (White)
    c = ws['E1']
    c.value = rec.fecha_fin
    c.number_format = 'dd/mm/yyyy'
    c.font = Font(bold=True, size=16)
    c.protection = CellProtection(locked=False)
    c.fill = white_fill

    # D2: ULTIMA RECAUDACION (Size 12)
    ws['D2'] = "ULTIMA RECAUDACION"
    ws['D2'].font = header_font
    
    # E2: Start Date (Size 11/Default, Editable -> White)
    c = ws['E2']
    c.value = rec.fecha_inicio
    c.number_format = 'dd/mm/yyyy'
    c.font = Font(bold=True) # Default size seems to be 11
    c.protection = CellProtection(locked=False)
    c.fill = white_fill
    
    # F2: DIAS (Size 12)
    ws['F2'] = "DIAS"
    ws['F2'].font = header_font
    
    # G2: Formula (Size 12)
    ws['G2'] = "=E1-E2"
    ws['G2'].font = header_font

    # D3: VERSION (Size 12)
    ws['D3'] = "VERSION"
    ws['D3'].font = header_font

    # E3: Version Value (1.0)
    c = ws['E3']
    c.value = "1.0"
    c.font = Font(bold=True)
    c.alignment = Alignment(horizontal='center') # Center align like dates usually are? Or left? Let's keep default or consistent. Dates were center/right? E1/E2 didn't specify alignment, so default.
    # Version should be static/read-only for machine process, so LOCKED (default is locked if sheet is protected).
    # Since we aren't setting protection=False, it will be locked.
    # And we want it Gray (default from base loop) or White? The plan said "Locked/Gray".
    # Base loop sets A1:H{t_row} to Gray. So it will be Gray by default.
    # No extra code needed for Gray+Locked.

    # --- Summary Section (Rows 3-10) ---
    # Specs from Plantilla02 Analysis
    # B3: Total Recaudado = SUM(B_data)+SUM(C_data) (Retirada + Cajon)
    # B4: Pagos Manuales = SUM(D_data)
    # B5: Ajustes = SUM(E_data)
    
    summary_data = [
        (3, "TOTAL RECAUDADO", f"=SUM(B{start_row}:B{end_row})+SUM(C{start_row}:C{end_row})", True, 14, "009900"),
        (4, "PAGOS MANUALES", f"=SUM(D{start_row}:D{end_row})", True, 14, "FF0000"),
        (5, "AJUSTES", f"=SUM(E{start_row}:E{end_row})", True, None, None),
        (6, "TOTAL TASAS", rec.total_tasas or 0, False, None, "FF0000"), # Editable -> White
        (7, "SUBTOTAL", "=B3-B4+B5-B6", True, None, None),
        (8, "DEPÓSITOS", rec.depositos or 0, False, None, None), # Editable -> White
        (9, "OTROS CONCEPTOS", rec.otros_conceptos or 0, False, None, None), # Editable -> White
        (10, "TOTAL:", "=B7+B8+B9", True, 16, None) # "TOTAL:" with colon
    ]

    for row_idx, label, val, is_locked, size, color in summary_data:
        # Label (Column A)
        ws.cell(row=row_idx, column=1, value=label).font = Font(bold=True, size=size if size else 11)
        
        # Value (Column B)
        c = ws.cell(row=row_idx, column=2, value=val)
        c.number_format = money_fmt
        
        # Font format
        f_size = size if size else 11
        f_color = color if color else "000000"
        c.font = Font(bold=True, size=f_size, color=f_color)
        
        # Locking & Color
        if not is_locked:
            c.protection = CellProtection(locked=False)
            c.fill = white_fill
        # Locked cells stay Gray from base loop

    # Extra Summary Fields (Row 10)
    # C10: LOCAL (50%)
    c = ws['C10']
    c.value = "LOCAL (50%)"
    c.font = Font(bold=True, size=16)
    
    # D10: Formula
    c = ws['D10']
    c.value = "=B10*0.5"
    c.number_format = money_fmt
    c.font = Font(bold=True, size=16)
    
    # E10: UORSA (50%)
    c = ws['E10']
    c.value = "UORSA (50%)"
    c.font = Font(bold=True, size=16)
    
    # F10: Formula
    c = ws['F10']
    c.value = "=B10*0.5"    
    c.number_format = money_fmt
    c.font = Font(bold=True, size=16)

    # --- Table Header (Row 12) ---
    headers_table = ["MÁQUINA", "RETIRADA EFECTIVO", "CAJÓN", "PAGO MANUAL", "AJUSTE", "TOTAL BRUTO", "TASA ESTIMADA", "TOTAL NETO"]
    for i, h in enumerate(headers_table, 1):
        c = ws.cell(row=12, column=i, value=h)
        c.font = bold_font
        c.border = Border(bottom=Side(style='thin'))
        c.alignment = Alignment(horizontal='center')
        # Fill is Gray from base loop
        
    ws.freeze_panes = "A13"
    
    ws.column_dimensions['A'].width = 31
    from openpyxl.utils import get_column_letter
    for col in range(2, 9):
        ws.column_dimensions[get_column_letter(col)].width = 18

    # --- Data (Row 13+) ---
    # Tenuous Border Style (Hair or Thin Gray)
    tenuous_side = Side(style='hair', color='888888')
    thick_side = Side(style='double', color='000000')
    
    count = len(data)
    for r_idx, d in enumerate(data, start=start_row):
        # Determine Block Boundaries
        idx = r_idx - start_row
        current_name = d["raw_name"]
        prev_name = data[idx-1]["raw_name"] if idx > 0 else None
        next_name = data[idx+1]["raw_name"] if idx < count - 1 else None
        
        is_top = (current_name != prev_name)
        is_bottom = (current_name != next_name)
        
        # Helper to get border for a cell
        def get_border(col_idx):
            top = thick_side if is_top else tenuous_side
            bottom = thick_side if is_bottom else tenuous_side
            left = tenuous_side
            right = tenuous_side
            return Border(top=top, bottom=bottom, left=left, right=right)

        # A: Machine (Gray)
        c = ws.cell(row=r_idx, column=1, value=d["MAQUINA"])
        c.font = bold_font
        c.border = get_border(1)
        
        # B, C, D, E: Editable -> White
        for col_idx, key in [(2, "RETIRADA"), (3, "CAJON"), (4, "PAGO MANUAL"), (5, "AJUSTE")]:
            c = ws.cell(row=r_idx, column=col_idx, value=d[key])
            c.number_format = money_fmt
            c.protection = CellProtection(locked=False)
            c.fill = white_fill
            c.border = get_border(col_idx)
            
            # D (Pago Manual) -> Static Red Text
            if col_idx == 4:
                c.font = red_static_font
        
        # F (Total Bruto): Formula matches Plantilla02 (=B+C-D+E)
        c = ws.cell(row=r_idx, column=6, value=f"=B{r_idx}+C{r_idx}-D{r_idx}+E{r_idx}")
        c.number_format = money_fmt
        c.border = get_border(6)
        
        # G (Tasa Est): Static Red Value. Locked -> Gray
        c = ws.cell(row=r_idx, column=7, value=d["TASA_EST"])
        c.number_format = money_fmt
        c.font = red_static_font
        c.border = get_border(7)
        # Gray fill default
        
        # H (Total Neto): Formula matches Plantilla02 (=F-G)
        c = ws.cell(row=r_idx, column=8, value=f"=F{r_idx}-G{r_idx}")
        c.number_format = money_fmt
        c.border = get_border(8)

    # --- Totals Row ---
    ws.cell(row=t_row, column=1, value="TOTAL").font = Font(bold=True, size=11)
    
    cols = ['B', 'C', 'D', 'E', 'F', 'G', 'H']
    for i, col in enumerate(cols, 2):
        c = ws.cell(row=t_row, column=i)
        c.value = f"=SUM({col}{start_row}:{col}{end_row})"
        c.number_format = money_fmt
        c.border = Border(top=Side(style='thin', color="000000")) 
        
        # Red static for D and G
        if i == 4 or i == 7:
            c.font = Font(bold=True, size=11, color="FF0000")
        else:
            c.font = Font(bold=True, size=11)
            
    # --- Conditional Formatting ---
    # Standard Rule (Green > 0, Red < 0)
    # Exclude D (4) and G (7).
    # Ranges: B{s}:C{t}, E{s}:F{t}, H{s}:H{t}
    
    range_BC = f"B{start_row}:C{t_row}"
    range_EF = f"E{start_row}:F{t_row}"
    range_H  = f"H{start_row}:H{t_row}"
    
    # Summary Cells: B5, B7, B8, B9, B10, D10, F10
    standard_cells = "B5 B7 B8 B9 B10 D10 F10"
    
    standard_ranges = f"{range_BC} {range_EF} {range_H} {standard_cells}"
    
    green_font = Font(color="009900", bold=True)
    rule_green = CellIsRule(operator='greaterThan', formula=['0'], stopIfTrue=False, font=green_font)
    ws.conditional_formatting.add(standard_ranges, rule_green)
    
    red_font = Font(color="FF0000", bold=True)
    rule_red = CellIsRule(operator='lessThan', formula=['0'], stopIfTrue=False, font=red_font)
    ws.conditional_formatting.add(standard_ranges, rule_red)
    
    wb.save(output)
    output.seek(0)
    
    # Filename: NOMBRE_SALON_AAAAMMDD_Recaudacion.xlsx
    salon_name = rec.salon.nombre.replace(" ", "_") if rec.salon and rec.salon.nombre else "Salon"
    date_str = rec.fecha_fin.strftime('%Y%m%d')
    filename = f"{salon_name}_{date_str}_Recaudacion.xlsx"
    headers = {
        'Content-Disposition': f'attachment; filename="{filename}"',
        'Access-Control-Expose-Headers': 'Content-Disposition' # Ensure regex sees it
    }
    return Response(content=output.getvalue(), media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', headers=headers)
