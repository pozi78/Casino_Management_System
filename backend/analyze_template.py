
import openpyxl
from openpyxl.styles import PatternFill

def analyze_template():
    try:
        wb = openpyxl.load_workbook('/app/Plantilla01.xlsx')
        ws = wb.active
        
        print(f"Sheet Title: {ws.title}")
        print(f"Max Row: {ws.max_row}, Max Col: {ws.max_column}")
        
        print("\n--- Headers & Content ---")
        for row in ws.iter_rows(min_row=1, max_row=15, max_col=10):
            row_data = []
            for cell in row:
                val = str(cell.value).strip() if cell.value is not None else ""
                row_data.append(f"{cell.coordinate}:{val}")
            print(" | ".join(row_data))

        print("\n--- Yellow Cells (To be Locked) ---")
        # Standard Yellow is FFFF00 or generic 'Yellow' color index. Openpyxl returns hex/theme.
        # We'll just look for non-none Fills.
        count = 0
        for row in ws.iter_rows(min_row=1, max_row=30, max_col=15):
            for cell in row:
                if cell.fill and cell.fill.start_color:
                    # Check for yellow-ish colors. Often FFFF00 or theme colors.
                    # Just print any color found to identify the "Yellow" code.
                    if cell.fill.fill_type == 'solid':
                         # Usually 'FFFF0000' is red, 'FFFF00' is yellow? 
                         # Let's print the RGB or Theme value
                         color = cell.fill.start_color
                         # If it looks like yellow (FFFF00), log it.
                         c_val = str(color.rgb) if color.type == 'rgb' else f"Theme:{color.theme},Tint:{color.tint}"
                         if 'FFFF00' in c_val or 'yellow' in str(c_val).lower(): 
                             print(f"Yellow/Locked Cell: {cell.coordinate} (Color: {c_val})")
                             count += 1
                         elif color.rgb != '00000000' and color.rgb is not None:
                             print(f"Other Colored Cell: {cell.coordinate} (Color: {c_val})")
        print(f"Total Yellow Candidates: {count}")

    except Exception as e:
        print(f"Error: {e}")

if __name__ == "__main__":
    analyze_template()
