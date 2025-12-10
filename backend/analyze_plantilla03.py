import openpyxl
from openpyxl.styles import PatternFill

def get_rgb(fill):
    if fill.start_color.type == 'rgb':
        return fill.start_color.rgb
    if fill.start_color.theme is not None:
        return f"Theme:{fill.start_color.theme}"
    return "None"

def get_border(cell):
    b = cell.border
    return f"L:{b.left.style if b.left else 'x'},R:{b.right.style if b.right else 'x'},T:{b.top.style if b.top else 'x'},B:{b.bottom.style if b.bottom else 'x'}"

wb = openpyxl.load_workbook('Plantilla03.xlsx')
ws = wb.active

print("--- Plantilla03.xlsx Analysis ---")
print(f"Title: {ws.title}")

# Analyze headers and structure
for row in [1, 2, 12, 13]: # Check key rows
    row_data = []
    for col in range(1, 10): # A-I
        c = ws.cell(row=row, column=col)
        val = str(c.value).strip() if c.value is not None else ""
        font_size = c.font.sz if c.font else "Def"
        border_info = get_border(c)
        
        if val or font_size != 11: # Filter interesting
           row_data.append(f"{c.coordinate}:['{val}', Sz:{font_size}, Bdr:{border_info}]")
            
    if row_data:
        print(f"Row {row}: " + ", ".join(row_data))
